from starkdata import StarkData
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from PIL import ImageTk
from time import strftime, localtime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import os


class App:
    def __init__(self, data):
        self.data = data
        self.root = tk.Tk()
        self.root.protocol("WM_DELETE_WINDOW", lambda: self.close())
        self.root.geometry("600x700")
        self.root.title("Stark App")
        self.root.iconbitmap("logo.ico")

        self.image = ImageTk.PhotoImage(file="iu.jpeg")
        tk.Label(self.root, image=self.image).place(relwidth=1, relheight=1)

        # Configures the search frame
        self.search_frame = tk.Frame(self.root, bg="red", bd=5)
        self.search_frame.place(relx=0.5,
                                rely=0.075,
                                relwidth=0.75,
                                relheight=0.15,
                                anchor="n")

        self.search_label = tk.Label(
            self.search_frame,
            text=
            "Digite o produto buscado e clique em 'Buscar'\n ou deixe em branco para obter todos os produtos",
            font="None 15")
        self.search_label.place(relheight=0.45, relwidth=1)

        self.search_bar = tk.Entry(self.search_frame, font="None 15")
        self.search_bar.place(rely=0.5, relwidth=0.69, relheight=0.5)

        self.search_button = tk.Button(self.search_frame,
                                       text="Buscar",
                                       font="None 15",
                                       command=lambda: self.fill_table())
        self.search_button.place(relx=0.7,
                                 rely=0.5,
                                 relwidth=0.30,
                                 relheight=0.5)

        # Configures the table of products
        self.table_frame = tk.Frame(self.root, bg="black", bd=5)
        self.table_frame.place(relx=0.5,
                               rely=0.25,
                               relwidth=0.75,
                               relheight=0.5,
                               anchor="n")

        self.table = ttk.Treeview(self.table_frame,
                                  columns=(1, 2, 3, 0),
                                  show="headings")
        self.table.place(relwidth=0.98, relheight=0.95)

        self.table.heading(1, text="Nome", anchor="w")
        self.table.heading(2, text="Estoque", anchor="w")
        self.table.heading(3, text="Preço", anchor="w")

        self.table.column(0, width=0, stretch=False)
        self.table.column(1, anchor="w", width=200, minwidth=200)
        self.table.column(2, anchor="w", width=40)
        self.table.column(3, anchor="w", width=40)

        self.table.tag_configure('greenrow',
                                 background="#02fb5e",
                                 font="None 12")
        self.table.tag_configure('redrow',
                                 background="#df1019",
                                 font="None 12")

        self.scroll_bar = ttk.Scrollbar(self.table_frame,
                                        orient="vertical",
                                        command=self.table.yview)
        self.scroll_bar.place(relx=0.98, relheight=0.95, relwidth=0.02)

        self.table.configure(yscrollcommand=self.scroll_bar.set)

        self.style = ttk.Style()
        self.style.theme_use("default")
        self.style.configure("Treeview",
                             background="#D3D3D3",
                             foreground="black",
                             rowheight=25,
                             fieldbackground="#D3D3D3")
        self.style.configure("Treeview.Heading", font=(None, 15))
        self.style.map("Treeview", background=[('selected', 'blue')])

        tk.Button(self.table_frame,
                  text="+",
                  font="None 12",
                  command=lambda: self.add_product()).place(relheight=0.05,
                                                            relwidth=1,
                                                            rely=0.95)

        # Configures the action buttons
        self.buttons_frame = tk.Frame(self.root, bg="red", bd=5)
        self.buttons_frame.place(relx=0.5,
                                 rely=0.8,
                                 relheight=0.15,
                                 relwidth=0.75,
                                 anchor="n")

        tk.Button(self.buttons_frame,
                  text="Alterar preço",
                  font="None 12",
                  command=lambda: self.change_price()).place(relheight=0.45,
                                                             relwidth=0.3,
                                                             relx=0.015)

        tk.Button(self.buttons_frame,
                  text="Alterar estoque",
                  font="None 12",
                  command=lambda: self.change_stock()).place(relheight=0.45,
                                                             relwidth=0.3,
                                                             relx=0.35)

        tk.Button(self.buttons_frame,
                  text="Deletar produtos",
                  font="None 12",
                  command=lambda: self.delete_product()).place(relheight=0.45,
                                                               relwidth=0.3,
                                                               relx=0.685)

        tk.Button(self.buttons_frame,
                  text="Gerar Relatório",
                  font="None 12",
                  command=lambda: self.create_report()).place(relheight=0.45,
                                                              relwidth=0.4775,
                                                              rely=0.55,
                                                              relx=0.015)

        tk.Button(self.buttons_frame,
                  text="Gerar histórico",
                  font="None 12",
                  command=lambda: self.show_history()).place(relheight=0.45,
                                                             relwidth=0.4775,
                                                             rely=0.55,
                                                             relx=0.5075)

    def show_history(self):
        """
        Creates a window to ask for the desired time.
        Then creates a file to show the history
        """
        history_window = tk.Toplevel(self.root)
        history_window.geometry("500x200")
        history_window.resizable(False, False)

        tk.Label(
            history_window,
            text=
            "Insira a data para produzir o histórico:\nformato: 'dia-mês-ano'",
            font="None 20 bold").place(relheight=0.3, relwidth=1, rely=0.05)
        tk.Label(history_window,
                 text="Data inicial:",
                 font="None 15 bold",
                 anchor="w").place(relheight=0.15,
                                   relwidth=0.27,
                                   rely=0.4,
                                   relx=0.05)
        start_date_entry = tk.Entry(history_window, font="None 12")
        start_date_entry.place(relheight=0.15,
                               relwidth=0.63,
                               rely=0.4,
                               relx=0.32)
        tk.Label(history_window,
                 text="Data final:",
                 font="None 15 bold",
                 anchor="w").place(relheight=0.15,
                                   relwidth=0.27,
                                   rely=0.55,
                                   relx=0.05)
        end_date_entry = tk.Entry(history_window, font="None 12")
        end_date_entry.place(relheight=0.15,
                             relwidth=0.63,
                             rely=0.55,
                             relx=0.32)

        tk.Button(history_window,
                  text="Cancelar",
                  fg="#df1019",
                  font="None 12",
                  command=lambda: history_window.destroy()).place(
                      rely=0.825, relx=0.06, relheight=0.1, relwidth=0.4)
        tk.Button(history_window,
                  text="Gerar",
                  font="None 12",
                  command=lambda: create_history()).place(rely=0.825,
                                                          relx=0.54,
                                                          relheight=0.1,
                                                          relwidth=0.4)

        def create_history():
            """
            Creates the history file
            """
            # Checks if the user input was a valid date
            start_date = start_date_entry.get()
            end_date = end_date_entry.get()

            try:
                list(map(int, start_date.split("-")))

                if len(start_date.split("-")) != 3:
                    raise ValueError
            except ValueError:
                self.raise_message("Data incial inválida!")
                return

            try:
                list(map(int, end_date.split("-")))

                if len(end_date.split("-")) != 3:
                    raise ValueError
            except ValueError:
                self.raise_message("Data final inválida")

            # Creates dates with desired format
            formatted_start_date = "-".join(
                reversed([item.zfill(2) for item in start_date.split("-")]))
            formatted_end_date = "-".join(
                reversed([item.zfill(2) for item in end_date.split("-")]))

            work_book = openpyxl.Workbook()
            sheet = work_book.active

            sheet.cell(row=1, column=1).value = "Descrição"
            sheet.cell(row=1, column=2).value = "Data"
            sheet.cell(row=1, column=3).value = "Hora"

            # Fills the rows with the actions in history
            for row, action in enumerate(
                    self.data.read_history(formatted_start_date,
                                           formatted_end_date)):
                sheet.cell(row=row + 2, column=1).value = action['description']
                sheet.cell(row=row + 2, column=2).value = "-".join(
                    reversed(action['date'].split("-")))
                sheet.cell(row=row + 2, column=3).value = action['time']

                if "vendidas" in action['description']:
                    greenFill = PatternFill(start_color='90EE90',
                                            end_color='90EE90',
                                            fill_type='solid')
                    sheet.cell(row=row + 2, column=1).fill = greenFill
                    sheet.cell(row=row + 2, column=2).fill = greenFill
                    sheet.cell(row=row + 2, column=3).fill = greenFill
                elif "dadas" in action['description']:
                    blueFill = PatternFill(start_color='19B2FF',
                                           end_color='19B2FF',
                                           fill_type='solid')
                    sheet.cell(row=row + 2, column=1).fill = blueFill
                    sheet.cell(row=row + 2, column=2).fill = blueFill
                    sheet.cell(row=row + 2, column=3).fill = blueFill
                elif "adicionadas" in action['description']:
                    orangeFill = PatternFill(start_color='E67300',
                                             end_color='E67300',
                                             fill_type='solid')
                    sheet.cell(row=row + 2, column=1).fill = orangeFill
                    sheet.cell(row=row + 2, column=2).fill = orangeFill
                    sheet.cell(row=row + 2, column=3).fill = orangeFill
                elif "alterado" in action['description']:
                    yellowFill = PatternFill(start_color='F2EA00',
                                             end_color='F2EA00',
                                             fill_type='solid')
                    sheet.cell(row=row + 2, column=1).fill = yellowFill
                    sheet.cell(row=row + 2, column=2).fill = yellowFill
                    sheet.cell(row=row + 2, column=3).fill = yellowFill

            sheet.column_dimensions[get_column_letter(1)].width = 100
            sheet.column_dimensions[get_column_letter(2)].width = 15
            sheet.column_dimensions[get_column_letter(3)].width = 15

            # Asks for the directory where the report will be in and saves it
            filepath = filedialog.askdirectory()
            filename = f"Histórico_{start_date}_{end_date}.xlsx"
            work_book.save(os.path.join(filepath, filename))

            history_window.destroy()

    def create_report(self):
        """
        Creates a report of all the products stock in the day
        Writes it in an excel file
        """

        work_book = openpyxl.Workbook()
        sheet = work_book.active

        sheet.cell(row=1, column=1).value = "Produto"
        sheet.cell(row=1, column=2).value = "Estoque"
        sheet.cell(row=1, column=3).value = "Preço"

        # Fills the rows with the products
        # Colors the row red if the product is not in stock
        for row, product in enumerate(self.data.search_product("")):
            sheet.cell(row=row + 2, column=1).value = product[
                'name'] + " " + product['color'] + " " + product['size']
            sheet.cell(row=row + 2, column=2).value = product['in_stock']
            sheet.cell(row=row + 2, column=3).value = product['price']
            sheet.cell(row=row + 2, column=3).number_format = 'R$0.00'

            if product['in_stock'] == 0:
                redFill = PatternFill(start_color='FFCCCB',
                                      end_color='FFCCCB',
                                      fill_type='solid')
                sheet.cell(row=row + 2, column=1).fill = redFill
                sheet.cell(row=row + 2, column=2).fill = redFill
                sheet.cell(row=row + 2, column=3).fill = redFill

        sheet.column_dimensions[get_column_letter(1)].width = 40
        sheet.column_dimensions[get_column_letter(2)].width = 15
        sheet.column_dimensions[get_column_letter(3)].width = 15

        # Asks for the directory where the report will be in and saves it
        filepath = filedialog.askdirectory()
        filename = strftime("Relatório_%d-%m-%Y.xlsx", localtime())
        work_book.save(os.path.join(filepath, filename))

    def delete_product(self):
        """
        Delete the selected product in the table.
        """
        # Checks if there is a product selected.
        # If there isn't, then the button should do nothing
        selected_products = list(map(int, self.table.selection()))
        amount_items = len(selected_products)
        if amount_items == 0:
            self.raise_message("Selecione pelo menos um produto para deletar!")
            return

        # Creates a confirm window
        confirm_window = tk.Toplevel(self.root)
        confirm_window.geometry("500x100")
        confirm_window.resizable(False, False)
        confirm_window.title("Confirmação")

        tk.Label(
            confirm_window,
            text=
            f"Tem certeza de que deseja apagar {amount_items} produtos?\nEssa ação não pode ser desfeita.",
            font="None 15 bold").place(relheight=0.5, relwidth=1)
        tk.Button(confirm_window,
                  text="Sim",
                  command=lambda: confirm(),
                  fg="red").place(relheight=0.45,
                                  relwidth=0.2,
                                  relx=0.6,
                                  rely=0.5)
        tk.Button(confirm_window,
                  text="Cancelar",
                  command=lambda: confirm_window.destroy()).place(
                      relheight=0.45, relwidth=0.2, relx=0.2, rely=0.5)

        def confirm():
            for product_id in selected_products:
                self.data.delete_product(product_id)
            self.fill_table()
            confirm_window.destroy()

    def add_product(self):
        """
        Creates a window that lets the user add a new product
        """
        product_window = tk.Toplevel(self.root)
        product_window.geometry("500x250")
        product_window.resizable(False, False)

        tk.Label(product_window,
                 text="Insira os dados do produto:",
                 font="None 20 bold").place(relheight=0.1,
                                            relwidth=1,
                                            rely=0.05)
        tk.Label(product_window, text="Nome:", font="None 15 bold",
                 anchor="w").place(relheight=0.1,
                                   relwidth=0.2,
                                   rely=0.2,
                                   relx=0.05)
        tk.Label(product_window, text="Cor:", font="None 15 bold",
                 anchor="w").place(relwidth=0.2,
                                   relheight=0.1,
                                   rely=0.3,
                                   relx=0.05)
        tk.Label(product_window,
                 text="Tamanho:",
                 font="None 15 bold",
                 anchor="w").place(relwidth=0.2,
                                   relheight=0.1,
                                   rely=0.4,
                                   relx=0.05)
        tk.Label(product_window,
                 text="Estoque:",
                 font="None 15 bold",
                 anchor="w").place(relwidth=0.2,
                                   relheight=0.1,
                                   rely=0.5,
                                   relx=0.05)
        tk.Label(product_window,
                 text="Preço(R$):",
                 font="None 15 bold",
                 anchor="w").place(relwidth=0.2,
                                   relheight=0.1,
                                   rely=0.6,
                                   relx=0.05)
        name_entry = tk.Entry(product_window, font="None 12")
        name_entry.place(relheight=0.1, relwidth=0.7, rely=0.2, relx=0.25)
        color_entry = tk.Entry(product_window, font="None 12")
        color_entry.place(relheight=0.1, relwidth=0.7, rely=0.3, relx=0.25)
        size_entry = tk.Entry(product_window, font="None 12")
        size_entry.place(relheight=0.1, relwidth=0.7, rely=0.4, relx=0.25)
        stock_entry = tk.Entry(product_window, font="None 12")
        stock_entry.place(relheight=0.1, relwidth=0.7, rely=0.5, relx=0.25)
        price_entry = tk.Entry(product_window, font="None 12")
        price_entry.place(relheight=0.1, relwidth=0.7, rely=0.6, relx=0.25)

        tk.Button(product_window,
                  text="Cancelar",
                  fg="#df1019",
                  font="None 12",
                  command=lambda: product_window.destroy()).place(
                      rely=0.775, relx=0.2, relheight=0.15, relwidth=0.2)
        tk.Button(product_window,
                  text="Adicionar",
                  fg="#008000",
                  font="None 12",
                  command=lambda: read_product()).place(rely=0.775,
                                                        relx=0.6,
                                                        relheight=0.15,
                                                        relwidth=0.2)

        def read_product():
            """
            Adds the product to database
            """
            name = name_entry.get()
            color = color_entry.get()
            size = size_entry.get()
            stock = stock_entry.get()
            price = price_entry.get()

            if stock == "":
                stock = 0
            else:
                try:
                    stock = int(stock)
                except ValueError:
                    self.raise_message(
                        "A quantidade em estoque tem que ser um número inteiro!"
                    )
                    return

            if price == "":
                price = 0
            else:
                try:
                    price = float(price)
                except ValueError:
                    self.raise_message("O preço tem que ser um número real!")
                    return

            if name == "":
                self.raise_message("Bocó, você esqueceu de colocar o nome!")
                return

            if color == "":
                self.raise_message(
                    "Onde já se viu um produto sem cor? Preenche isso direito!"
                )
                return

            if size not in ["PP", "P", "M", "G", "GG", "unico"]:
                self.raise_message(
                    "Tamanho inválido! Aqui estão os possíveis tamanhos:\n PP, P, M, G, GG ou unico (sem assento)"
                )
                return

            self.data.add_product((name, color, size, stock, price))

            # Destroys the window
            product_window.destroy()

            # Deletes the content on search bar then fills the table
            self.search_bar.delete(0, len(self.search_bar.get()))
            self.fill_table()

    def change_stock(self):
        """
        Gives the option to the user to change the stocks of a product
        """
        try:
            product_id = int(self.table.focus())
        except ValueError:
            self.raise_message("Selecione um produto para alterar o estoque!")
            return

        stock_window = tk.Toplevel(self.root)
        stock_window.geometry("500x150")
        stock_window.resizable(False, False)

        tk.Label(stock_window,
                 text="Insira a quantidade e clique na ação:",
                 font="None 20 bold").place(relheight=0.3,
                                            relwidth=1,
                                            rely=0.05)
        tk.Label(stock_window,
                 text="Quantidade:",
                 font="None 15 bold",
                 anchor="w").place(relheight=0.2,
                                   relwidth=0.27,
                                   rely=0.4,
                                   relx=0.05)
        quantity_entry = tk.Entry(stock_window, font="None 12")
        quantity_entry.place(relheight=0.2, relwidth=0.63, rely=0.4, relx=0.32)

        tk.Button(stock_window,
                  text="Cancelar",
                  fg="#df1019",
                  font="None 12",
                  command=lambda: stock_window.destroy()).place(rely=0.725,
                                                                relx=0.04,
                                                                relheight=0.15,
                                                                relwidth=0.2)
        tk.Button(stock_window,
                  text="Adicionar",
                  font="None 12",
                  command=lambda: adjust_stock("add")).place(rely=0.725,
                                                             relx=0.28,
                                                             relheight=0.15,
                                                             relwidth=0.2)
        tk.Button(stock_window,
                  text="Remover",
                  font="None 12",
                  command=lambda: adjust_stock("remove")).place(rely=0.725,
                                                                relx=0.52,
                                                                relheight=0.15,
                                                                relwidth=0.2)
        tk.Button(stock_window,
                  text="Alterar",
                  font="None 12",
                  command=lambda: adjust_stock("change")).place(rely=0.725,
                                                                relx=0.76,
                                                                relheight=0.15,
                                                                relwidth=0.2)

        def adjust_stock(action):
            """
            Adjust the stock value for the item, depending on the action
            """
            # Checks if the quantity given is an integer
            try:
                quantity = int(quantity_entry.get())
            except ValueError:
                self.raise_message(
                    "Aiai, eu avisei que a quantidade tinha que ser um inteiro!"
                )
                return

            if quantity < 0:
                self.raise_message("A quantidade tem que ser não negativa!")
                return

            product = self.data.get_product(product_id)

            if action == "remove":
                if quantity > product['in_stock']:
                    self.raise_message(
                        "Tá querendo tirar mais do que tem?\nQuer moleza? Senta num pudim!"
                    )
                    return

                new_quantity = product['in_stock'] - quantity

                new_history_window = tk.Toplevel(self.root)
                new_history_window.geometry("700x150")
                new_history_window.resizable(False, False)

                tk.Label(new_history_window,
                         text="Insira para quem o produto foi dado/vendido:",
                         font="None 20 bold").place(relheight=0.3,
                                                    relwidth=1,
                                                    rely=0.05)
                tk.Label(new_history_window,
                         text="Nome:",
                         font="None 15 bold",
                         anchor="w").place(relheight=0.2,
                                           relwidth=0.27,
                                           rely=0.4,
                                           relx=0.05)
                name_entry = tk.Entry(new_history_window, font="None 12")
                name_entry.place(relheight=0.2,
                                 relwidth=0.63,
                                 rely=0.4,
                                 relx=0.32)

                tk.Button(new_history_window,
                          text="Dado",
                          fg="#19B2FF",
                          font="None 12",
                          command=lambda: add_history("dadas")).place(
                              rely=0.725,
                              relx=0.06,
                              relheight=0.15,
                              relwidth=0.4)
                tk.Button(new_history_window,
                          text="Vendido",
                          fg="#90EE90",
                          font="None 12",
                          command=lambda: add_history("vendidas")).place(
                              rely=0.725,
                              relx=0.54,
                              relheight=0.15,
                              relwidth=0.4)

                def add_history(action):
                    """
                    Adds the action to the history.
                    """
                    message_string = f"{quantity} unidades do produto {product['name']} {product['color']} {product['size']} foram {action} para {name_entry.get()}"

                    self.data.add_history(message_string)

                    new_history_window.destroy()

            elif action == "add":
                new_quantity = product['in_stock'] + quantity

                self.data.add_history(
                    f"{quantity} unidades do produto {product['name']} {product['color']} {product['size']} foram adicionadas"
                )
            elif action == "change":
                new_quantity = quantity

                self.data.add_history(
                    f"Estoque do produto {product['name']} {product['color']} {product['size']} foi alterado para {quantity}"
                )

            self.data.update_stock(product_id, new_quantity)

            stock_window.destroy()

            self.raise_message(
                f"O estoque de {product['name']} {product['color']} {product['size']} foi atualizado para {new_quantity}",
                "Mensagem")

            self.fill_table()

    def change_price(self):
        """
        Gives the option to the user to change the price of a product
        """
        try:
            product_id = int(self.table.focus())
        except ValueError:
            self.raise_message("Selecione um produto para alterar o preço!")
            return

        price_window = tk.Toplevel(self.root)
        price_window.geometry("500x150")
        price_window.resizable(False, False)

        tk.Label(price_window,
                 text="Insira o novo preço:",
                 font="None 20 bold").place(relheight=0.3,
                                            relwidth=1,
                                            rely=0.05)
        tk.Label(price_window, text="Preço:", font="None 15 bold",
                 anchor="w").place(relheight=0.2,
                                   relwidth=0.27,
                                   rely=0.4,
                                   relx=0.05)
        price_entry = tk.Entry(price_window, font="None 12")
        price_entry.place(relheight=0.2, relwidth=0.63, rely=0.4, relx=0.32)

        tk.Button(price_window,
                  text="Cancelar",
                  fg="#df1019",
                  font="None 12",
                  command=lambda: price_window.destroy()).place(rely=0.725,
                                                                relx=0.06,
                                                                relheight=0.15,
                                                                relwidth=0.4)
        tk.Button(price_window,
                  text="Alterar",
                  font="None 12",
                  command=lambda: adjust_price()).place(rely=0.725,
                                                        relx=0.54,
                                                        relheight=0.15,
                                                        relwidth=0.4)

        def adjust_price():
            """
            Adjust the price value for the item
            """
            # Checks if the quantity given is an integer
            try:
                new_price = float(price_entry.get().replace(",", "."))
            except ValueError:
                self.raise_message(
                    "Aiai, eu avisei que a quantidade tinha que ser um real!")
                return

            if new_price < 0:
                self.raise_message("O preço tem que ser não negativo!")
                return

            product = self.data.get_product(product_id)

            self.data.update_price(product_id, new_price)
            self.data.add_history(
                f"Preço do produto {product['name']} {product['color']} {product['size']} foi alterado de R${product['price']:.2f} para R${new_price:.2f}"
                .replace(".", ","))

            price_window.destroy()

            self.raise_message(
                f"O preço de {product['name']} {product['color']} {product['size']} foi atualizado para R$"
                + f"{new_price:.2f}".replace(".", ","), "Mensagem")

            self.fill_table()

    def raise_message(self, message, title="Alerta!"):
        """
        Creates a message window showing the message 'message'
        By default it shows a warning
        """
        message_window = tk.Toplevel(self.root)
        message_window.resizable(False, False)
        message_window.title(title)

        tk.Label(message_window, text=message,
                 font="None 20 bold").pack(padx=5)
        tk.Button(message_window,
                  text="Ok",
                  font="None 15",
                  fg="blue",
                  command=lambda: message_window.destroy()).pack(pady=5)

    def fill_table(self):
        """
        Fills the table with the products
        """
        name = self.search_bar.get()
        produtcs = self.data.search_product(name)

        # Clears table
        for item in self.table.get_children():
            self.table.delete(item)

        # Adds the products
        for product in produtcs:
            if product['in_stock']:
                self.table.insert(
                    "",
                    "end",
                    iid=product['id'],
                    values=
                    (f"{product['name']} {product['color']} {product['size']}",
                     product['in_stock'],
                     "R$" + f"{product['price']:.2f}".replace(".", ",")),
                    tags=('greenrow', ))
            else:
                self.table.insert(
                    "",
                    "end",
                    iid=product['id'],
                    values=
                    (f"{product['name']} {product['color']} {product['size']}",
                     product['in_stock'],
                     "R$" + f"{product['price']:.2f}".replace(".", ",")),
                    tags=('redrow', ))

    def close(self):
        self.data.close()
        self.root.destroy()

    def run(self):
        self.root.mainloop()


def main():
    data = StarkData("stark.db")

    app = App(data)

    app.run()


if __name__ == "__main__":
    main()
